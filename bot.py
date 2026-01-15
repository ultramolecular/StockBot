#-----------------------------------------------------------------------------------------#
# Author:             Josiah Valdez                                                       #
# Began Development:  February, 7, 2021                                                   #
#                                                                                         #
# This is a program that uses selenium to extract data from tradingview.com top gainers,  #
# calculate and store their prices and percent changes, and send notifications given the  #
# desired change from the user. It is in constant development, and is sometimes unstable. #
#-----------------------------------------------------------------------------------------#
from dotenv import load_dotenv
from datetime import datetime as dt
from datetime import timedelta
from millify import millify
from float_provider import FloatProvider
import openpyxl
import os
from pathlib import Path
from platform import system
from time import sleep
from typing import Callable, Optional
from selenium import webdriver
from selenium.common.exceptions import TimeoutException, StaleElementReferenceException, WebDriverException
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from signal_scorer import SignalFeatures, SignalScorer
from rich import box, style
from rich.console import Console
from rich.table import Table
from stock import Stock
from webdriver_manager.chrome import ChromeDriverManager

##############################################################################
#                           GLOBAL CONFIG                                    #
##############################################################################

env_path = Path(".") / ".env"
load_dotenv(dotenv_path=env_path)
OS = system()
LOGIN_ICON_CSS = "button[aria-label='Open user menu']"
EMAIL_CSS = ".emailButton-nKAw8Hvt"
CREDS_IDS = ("id_username", "id_password")
LOGIN_BUTTON_CSS = ".submitButton-LQwxK8Bm"
STOCK_LIST_CSS = "tr[class*='listRow']"
STOCK_CSS = "td:nth-child(1) a"
STOCK_PRICES_CSS = "td:nth-child(3)"
VOL_CSS = "td:nth-child(4)"
HEADER_TABS_CSS = "market-screener-header-columnset-tabs"
URL = "https://www.tradingview.com"
GAINS_URL = "https://www.tradingview.com/markets/stocks-usa/market-movers-gainers/"
EMAIL = os.getenv("EMAIL", "")
PASS = os.getenv("PASS", "")
SOUND = "swiftly.wav"
MARKET_MONDAY = 0
MARKET_FRIDAY = 4
OPEN_HR = int(os.getenv("OPEN_HR", ""))
CLOSE_HR = int(os.getenv("CLOSE_HR", ""))
HEADLESS = int(os.getenv("HEADLESS", "0"))
EOD_EXPORT_DIR = os.getenv("EOD_EXPORT_PATH", "")
EXPORT_PATH = Path(EOD_EXPORT_DIR) if EOD_EXPORT_DIR else None

##############################################################################
#                           HELPER FUNCTIONS                                 #
##############################################################################


def setup_webdriver() -> webdriver.Chrome:
    """
    Creates and configures the Selenium Chrome WebDriver instance.
    """
    options = webdriver.ChromeOptions()
    if OS == "Windows":
        # Suppress extraneous logging in Windows
        options.add_experimental_option("excludeSwitches", ["enable-logging"])

    profile_dir = os.getenv("TV_PROFILE_PATH", "")
    profile_name = os.getenv("TV_PROFILE_NAME", "Default")

    if profile_dir:
        options.add_argument(f"--user-data-dir={profile_dir}")
        options.add_argument(f"--profile-directory={profile_name}")

    if HEADLESS: 
        options.add_argument("--headless=new")
        options.add_argument("--window-size=1300,1044")

    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(options=options, service=service)

    driver.set_window_size(1300, 1044)
    driver.set_page_load_timeout(20)
    return driver


def is_recaptcha_visible(driver: webdriver.Chrome) -> bool:
    """
    Detect recaptcha iframe visibility and returns True if visible, False otherwise.
    """
    return driver.execute_script(
        """
        var iframe = document.querySelector("iframe[title='reCAPTCHA']");
        if (iframe) {
            var style = window.getComputedStyle(iframe);
            var display = style && style.display !== 'none'
                              && style.visibility !== 'hidden'
                              && style.opacity !== '0';
            var rect = iframe.getBoundingClientRect();
            var inViewport = rect.width > 0 && rect.height > 0;
            return display && inViewport;
        } else {
            return false;
        }
        """
    )


def play_sound(file: str):
    """
    Play sound for user notifications. Different approach for Windows vs. other OS.
    """
    if OS == "Windows":
        import winsound

        winsound.PlaySound(file, winsound.SND_ASYNC)
    else:
        from playsound import playsound

        playsound(file)


def get_next_day(now=None) -> tuple[dt, dt]:
    if now is None:
        now = dt.now()

    wday = now.weekday()
    start_date = now.replace(hour=OPEN_HR, minute=30, second=30)
    end_date = now.replace(hour=CLOSE_HR, minute=0, second=0)

    # If it's a weekend go to next Monday market open
    if wday >= 5:
        # Calc days until Monday
        days_ahead = (7 - wday) % 7
        next_mon = start_date + timedelta(days=days_ahead)
        # Update MARKET_CLOSE to account for the new date
        return next_mon, next_mon.replace(hour=CLOSE_HR, minute=0, second=0)
    
    # If a weekday and it's before market open and close, run then
    if now < start_date or now < end_date:
        return start_date, end_date
    else:
        # After market close (still a weekday), schedule next day
        next_day = start_date + timedelta(days=1)
        # Check if the next day is Saturday to skip to Monday
        if next_day.weekday() == 5:
            next_day += timedelta(days=2) 

        return next_day, next_day.replace(hour=CLOSE_HR, minute=0, second=0)


def get_user_params() -> tuple[float, float]:
    """
    Prompt user for refresh rate, desired percent change, and secondary threshold.
    Returns (ref_rate_desSeconds, pct_chg_des, pct_chg_after).
    """
    ref_rate_des = float(input("Enter refresh rate (minutes) desired: ")) * 60
    pct_chg_des = float(input("Enter percent change desired (y% format): "))

    return ref_rate_des, pct_chg_des


def is_logged_in(driver: webdriver.Chrome, timeout: int = 10) -> bool:
    """
    Login check:
    - Navigate to an auth-gated TradingView URL
    - If redirected to sign-in or see sign-in wall text, not logged in
    """
    gated_url = URL + "/u"

    driver.get(gated_url)

    WebDriverWait(driver, timeout).until(
        lambda d: d.execute_script("return document.readyState") == "complete"
    )

    cur = driver.current_url.lower()
    if "signin" in cur or "sign-in" in cur or "login" in cur:
        return False

    # Fallback: check for common logged-out wall text
    page_text = driver.execute_script(
            "return (document.body && document.body.innerText) ? document.body.innerText : ''")
    lowered = page_text.lower()

    # Keep these minimal so it doesn't false-positive
    if "sign in" in lowered and "password" in lowered:
        return False

    return True


def login_to_tradingview(driver: webdriver.Chrome, email: str, password: str):
    """
    Logs in to TradingView by clicking the login icon, selecting email, entering credentials.
    """
    login_icon = WebDriverWait(driver, 2).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, LOGIN_ICON_CSS))
    )
    login_icon.click()

    actions = ActionChains(driver)
    actions.move_to_element(login_icon).perform()
    actions.send_keys(Keys.ARROW_DOWN).send_keys(Keys.RETURN).perform()

    WebDriverWait(driver, 2).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, EMAIL_CSS))
    ).click()

    WebDriverWait(driver, 2).until(
        EC.visibility_of_element_located((By.ID, CREDS_IDS[0]))
    ).send_keys(email)

    sleep(0.5)
    WebDriverWait(driver, 2).until(
        EC.visibility_of_element_located((By.ID, CREDS_IDS[1]))
    ).send_keys(password)

    sleep(0.5)
    WebDriverWait(driver, 2).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, LOGIN_BUTTON_CSS))
    ).click()


def check_recaptcha(driver: webdriver.Chrome):
    """
    Detect reCAPTCHA and wait for it to be solved if present. Otherwise, proceed.
    """
    try:
        WebDriverWait(driver, 4).until(lambda drv: is_recaptcha_visible(drv))
        print("-" * 120, "\nreCAPTCHA detected, please solve to continue!")
        try:
            WebDriverWait(driver, 60).until(
                lambda drv: drv.execute_script(
                    "return document.getElementsByName('g-recaptcha-response')[0].value !== '';"
                )
            )
            print("reCAPTCHA solved! Continuing...")
            WebDriverWait(driver, 2).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, LOGIN_BUTTON_CSS))
            ).click()
        except TimeoutException:
            print("Timeout waiting for reCAPTCHA to be solved. Exiting.")
            driver.close()
            driver.quit()
            exit()
    except TimeoutException:
        print("No reCAPTCHA detected. Continuing...")


def wait_for_login(driver: webdriver.Chrome, timeout: int = 15):
    """
    Wait for the login dialog to be gone; this indicates TradingView
    has finished processing the login. Mainly for reCAPTCHA logins.
    """
    # Wait until username field is no longer visible
    WebDriverWait(driver, timeout).until(
            EC.invisibility_of_element_located((By.ID, CREDS_IDS[0]))
    )


def wait_for_table(driver: webdriver.Chrome, timeout=15):
    """
    Wait until the gainers table loads by confirming at least one row
    has a valid ticker and price.
    """
    WebDriverWait(driver, timeout).until(
        lambda d: d.execute_script(
            """
            const rows = document.querySelectorAll("tr[class*='listRow']");
            if (!rows || rows.length === 0) return false;

            const c = rows[0].querySelectorAll("td");
            if (!c || c.length < 3) return false;

            const ticker = (c[0].querySelector("a")?.innerText || c[0].innerText).trim();
            const price = c[2].innerText.trim().replace(/[^0-9.]/g, '');

            return ticker.length > 0 && price.length > 0;
            """
        )
    )


def header_fields(driver: webdriver.Chrome) -> list[str]:
    return driver.execute_script(
            """
        const ths = document.querySelectorAll("table thead th[data-field]");
        return Array.from(ths).map(th => th.getAttribute("data-field") || "");
        """
    )


def ensure_header_mode(
    driver: webdriver.Chrome,
    mode: str,
    *,
    timeout: int = 8,
    retries: int = 3,
) -> bool:
    """
    Attempts to switch to 'mode' tab and verify by header data-field.
    Returns True if verified, False otherwise. Never raises TimeoutException.
    """
    # Fast-path: already in that mode
    try:
        want = "RelativeVolume" if mode == "overview" else "RelativeStrengthIndex"
        fields = header_fields(driver)
        if any(want in f for f in fields):
            return True
    except Exception:
        pass

    for attempt in range(1, retries + 1):
        try:
            click_header_tab(driver, mode)
            wait_for_header_mode(driver, mode, timeout=timeout)
            return True

        except (TimeoutException, StaleElementReferenceException, WebDriverException) as e:
            # small backoff + tiny nudge scroll (TV sometimes re-renders on scroll)
            sleep(0.25 * attempt)
            try:
                driver.execute_script("window.scrollBy(0, 1); window.scrollBy(0, -1);")
            except Exception:
                pass

    return False


def wait_for_header_mode(
    driver: webdriver.Chrome,
    mode: str,
    timeout: int = 10,
) -> None:
    want = "RelativeVolume" if mode == "overview" else "RelativeStrengthIndex"

    def _ok(d):
        fields = header_fields(d)
        return any(want in f for f in fields)

    WebDriverWait(driver, timeout).until(_ok)


# NOTE: DEBUG FUNC
def debug_dump_header_data_fields(driver) -> list[str]:
    return driver.execute_script(
        """
        return Array.from(document.querySelectorAll("thead th[data-field]"))
            .map(th => th.getAttribute("data-field"));
        """
    )


def click_header_tab(driver: webdriver.Chrome, tab_id: str):
    """
    Clicks the given header tab on the top gainers table, overview/technicals.
    """
    container = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.ID, HEADER_TABS_CSS))
    )

    btn = container.find_element(By.ID, tab_id)
    if btn.get_attribute("aria-selected") == "true":
        return

    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
    driver.execute_script("arguments[0].click();", btn)

    def _selected(d):
        c = d.find_element(By.ID, HEADER_TABS_CSS)
        b = c.find_element(By.ID, tab_id)
        return b.get_attribute("aria-selected") == "true"

    WebDriverWait(driver, 10).until(_selected)


def scrape_overview(driver: webdriver.Chrome) -> list[dict]:
    """
    Scrape overview tab from TradingView: ticker, price, volume, rvol.
    """
    if not ensure_header_mode(driver, "overview", timeout=6, retries=3):
        return []

    rows = driver.execute_script(
        """
        const rowCss = arguments[0];
        return Array.from(document.querySelectorAll(rowCss))
            .slice(0, 100)
            .map(r => {
                const c = r.querySelectorAll("td");
                if (c.length < 5) return null;

                const ticker =
                    (c[0].querySelector("a")?.innerText || c[0].innerText)
                        .trim();

                const priceText = c[2].innerText.trim().replace(/[^0-9.]/g, '');
                const price = parseFloat(priceText);
                const vol = c[3].innerText.trim();

                const rvolText = c[4].innerText.trim().replace(/[^0-9.]/g, '');
                const rvol = parseFloat(rvolText);

                if (!ticker || Number.isNaN(price)) return null;
                return { ticker, price, vol, rvol };
            })
            .filter(r => r !== null);
        """,
        STOCK_LIST_CSS,
    )
    return rows

def scrape_technicals(driver: webdriver.Chrome) -> dict[str, float]:
    """
    Scrape technicals tab for RSI.
    """
    if not ensure_header_mode(driver, "technicals", timeout=6, retries=3):
        # skip RSI this cycle; don't crash!
        return {}

    rows = driver.execute_script(
        """
        const rowCss = arguments[0];
        const result = {};
        Array.from(document.querySelectorAll(rowCss))
            .slice(0, 100)
            .forEach(r => {
                const c = r.querySelectorAll("td");
                if (c.length < 5) return;
                const ticker =
                    (c[0].querySelector("a")?.innerText || c[0].innerText)
                        .trim();
                const rsiText = c[4].innerText.trim().replace(/[^0-9.]/g, '');
                const rsi = parseFloat(rsiText);
                if (!ticker || Number.isNaN(rsi)) return;
                result[ticker] = rsi;
            });
        return result;
        """,
        STOCK_LIST_CSS,
    )
    return rows


def in_gainers(gainers: list[Stock], filt: Callable[[Stock], bool]) -> Optional[Stock]:
    """
    Searches for a stock of interest in the list and returns it if it exists.
    """
    for stock in gainers:
        if filt(stock):
            return stock
    return None


def process_stocks(
    gainers: list[Stock],
    float_prov: FloatProvider,
    alerted: list[Stock],
    new_data: list[tuple[str, float, str, float, float]],
    pct_chg_des: float,
) -> bool:
    """
    Update each Stock or create new ones based on the newly scraped data.
    Checks if user criteria is met, plays sounds, etc.
    Returns True if anything changed, otherwise False.
    """
    changed = False
    for stk_name, price, vol, rvol, rsi in new_data:
        stk = in_gainers(gainers, lambda s: s.get_ticker() == stk_name)
        if stk:
            new_abs_pct_chg = stk.get_new_abs(price)
            stk.set_new_after(price)

            # Keep last technicals up to date
            stk.update_technicals(price, vol, dt.now(), rvol, rsi)

            if new_abs_pct_chg != stk.get_abs():
                changed = True
                stk.set_abs(new_abs_pct_chg)
                stk.set_price(price)

            # check user criteria
            if stk.get_abs() >= pct_chg_des and not stk.has_met_crit():
                # Get float for this stock + cache it
                if stk.get_float_shares() is None:
                    stk.set_float_shares(float_prov.get_float_shares(stk.get_ticker()))
                # Recompute vol/float now that we have a float value known
                stk.update_technicals(price, vol, dt.now(), rvol, rsi)
                # Build features for scoring
                vol_shares = Stock.parse_volume_to_shares(vol) or 0.0
                feats = SignalFeatures(
                    ticker=stk.get_ticker(),
                    price=price,
                    abs_pct=stk.get_abs(),
                    volume=vol_shares,
                    float_shares=stk.float_shares,
                    rvol=stk.last_rvol,
                    rsi=stk.last_rsi,
                    vol_float_ratio=stk.last_vol_float_ratio,
                    time=dt.now(),
                )
                score_obj = SignalScorer.score(feats)
                stk.snapshot_crit_technicals(score_obj.score, score_obj.tier)

                # Log simple line for now
                print(
                    f"\n{stk.get_ticker()} +{stk.get_abs():.2f}% "
                    f"| Score: {score_obj.score:+d} ({score_obj.tier})"
                    f"| {dt.now().strftime("%H:%M:%S")}"
                )

                stk.did_meet_crit()
                stk.set_base_price(price)
                alerted.append(stk)
                show_alert_watchlist(alerted[:10])
                play_sound(SOUND)
        else:
            # brand new stock
            new_stock = Stock(stk_name, price, vol, rvol, rsi, dt.now())
            gainers.append(new_stock)
            changed = True

    return changed


def colorize_pct(val: float) -> str:
    """
    Return a Rich color markup string for the percentage, e.g. [green]2.53%[/green].
    Positive => green, Negative => red, Zero => white.
    """
    if val > 0:
        color = "green"
    elif val < 0:
        color = "red"
    else:
        color = "white"
    return f"[{color}]{val:6.2f}%[/{color}]"


def safe_pct(stk_age: int, val: float, min_age: int) -> str:
    """
    Return a colorized percent if the stock's age >= min_age, else '--'.
    This avoids showing intervals not yet 'lived'.
    """
    if stk_age >= min_age:
        return colorize_pct(val)
    else:
        return "--"


def style_for_score(score: int) -> str:
    if score >= 11:
        return "bold grey50 on bright_green"
    elif score >= 8:
        return "bold black on chartreuse3"
    elif score >= 5:
        return "bold black on yellow3"
    else:
        return "bold white on red3"


def style_for_rsi(rsi: Optional[float]) -> str:
    if rsi is None:
        return "white on grey23"
    if 60 <= rsi <= 75:
        return "black on bright_green"
    elif 50 <= rsi < 60 or 75 < rsi <= 85:
        return "bold black on yellow3"
    else:
        return "bold white on red3"


def style_for_rvol(rvol: Optional[float]) -> str:
    if rvol is None:
        return "white on grey23"
    if rvol >= 10:
        return "bold black on bright_green"
    elif rvol >= 5:
        return "bold black on yellow3"
    else:
        return "bold white on red3"


def style_for_vol(vol_str: Optional[str], float_shares: Optional[float] = None) -> str:
    """
    Heatmap style for volume.
    If float_shares is provided, color based on vol/float ratio (participation).
    Otherwise, color based on absolute volume bands.
    """
    if not vol_str:
        return "white on grey23"

    vol_shares = Stock.parse_volume_to_shares(vol_str)
    if not vol_shares or vol_shares <= 0:
        return "white on grey23"

    # Prefer vol/float if float is known
    if float_shares and float_shares > 0:
        vf = vol_shares / float_shares  # float rotation

        # Conservative bands: just a quick "is this meaningful yet?"
        if vf >= 1.0:
            return "bold black on bright_green"
        elif vf >= 0.5:
            return "bold black on yellow3"
        else:
            return "bold white on red3"

    # Absolute volume bands (fallback)
    if vol_shares >= 5_000_000:
        return "bold black on bright_green"
    elif vol_shares >= 1_000_000:
        return "bold black on green3"
    elif vol_shares >= 500_000:
        return "bold black on yellow3"
    elif vol_shares >= 100_000:
        return "bold white on dark_orange"
    else:
        return "bold white on red3"


def style_for_float(float_shares: Optional[float]) -> str:
    if not float_shares:
        return "white on grey23"

    m = float_shares / 1_000_000
    if m < 1:
        return "bold black on bright_green"
    elif m < 3:
        return "bold black on chartreuse3"
    elif m < 10:
        return "bold black on yellow3"
    elif m < 20:
        return "bold white on dark_orange"
    else:
        return "bold white on red3"


def style_for_fr(fr: Optional[float]) -> str:
    if fr is None:
        return "white on grey23"
    if fr >= 1.0:
        return "bold grey50 on bright_green"
    elif fr >= 0.5:
        return "bold black on yellow3"
    else:
        return "bold white on red3"


def show_alert_watchlist(stocks: list[Stock]) -> None:
    """
    Render up to 10 alerted stocks in a table/card layout for user to see.
    """
    if not stocks:
        return
    
    console = Console()
    table = Table(
        show_header=False,
        box=box.SIMPLE_HEAVY,
        pad_edge=False,
        expand=False,
    )

    for _ in range(7):
        table.add_column(no_wrap=True)

    for s in stocks[:10]:
        # --- basic time & growth metrics ---
        ticker = s.get_ticker()
        score = s.get_crit_score() or 0
        tier = s.get_crit_tier() or "-"

        # SPOTTED
        time_spotted = getattr(s, "TIME_ENTERED", None)
        time_spotted_str = (
            time_spotted.strftime("%H:%M:%S") if time_spotted else "n/a"
        )

        # ALERTED
        crit_time = s.get_crit_time()
        time_alerted_str = (
            crit_time.strftime("%H:%M:%S") if crit_time else "n/a"
        )

        # PEAK
        peak_time = s.get_time_max_price()
        time_peaked_str = (
            peak_time.strftime("%H:%M:%S") if peak_time else "n/a"
        )

        time_to_peak = s.get_time_peak_alert()
        time_to_peak_str = f"{time_to_peak} min" if time_to_peak is not None else "n/a"

        # Growths
        g_spot_to_peak = s.get_peak_change_spot()
        g_alert_to_peak = s.get_peak_change()

        # Float & FR
        float_shares = s.get_float_shares()
        float_m = float_shares / 1_000_000 if float_shares else None
        float_str = (
                f"[{style_for_float(float_shares)}]{float_m:.2f}M[/]"
                if float_m else "n/a"
        )

        fr_alert = s.get_crit_vol_float_ratio()
        fr_alert_str = f"{fr_alert:.2f}x" if fr_alert is not None else "n/a"

        # FR at peak: use volume_at_max_price / float
        fr_peak = None
        if float_shares and float_shares > 0:
            vol_peak_shares = Stock.parse_volume_to_shares(s.get_vol_at_max_price())
            if vol_peak_shares:
                fr_peak = vol_peak_shares / float_shares
        fr_peak_str = f"{fr_peak:.2f}x" if fr_peak is not None else "n/a"

        # RV / RSI snapshots
        # Spot RV/RSI
        rv_spot = s.get_og_rvol()
        rsi_spot = s.get_og_rsi()
        # Alert
        rv_alert = s.get_crit_rvol()
        rsi_alert = s.get_crit_rsi()
        # Peak
        rv_peak = s.get_peak_rvol()
        rsi_peak = s.get_peak_rsi()

        # Prices & volumes
        price_spot = s.get_og_price()
        vol_spot = s.get_og_vol()

        price_alert = s.get_crit_price()
        vol_alert = s.get_crit_vol()

        price_peak = s.get_max_price()
        vol_peak = s.get_vol_at_max_price()

        price_spot_str = f"{price_spot:.2f}" if price_spot is not None else "n/a"
        price_alert_str = f"{price_alert:.2f}" if price_alert is not None else "n/a"
        price_peak_str = f"{price_peak:.2f}" if price_peak is not None else "n/a"

        vol_spot_str = vol_spot or "n/a"
        vol_alert_str = vol_alert or "n/a"
        vol_peak_str = vol_peak or "n/a"

        # Row 1: Tick+Score | TimeSpot | Float | RV | RSI | Price | Vol
        row1_col1 = (
            f"[{style_for_score(score)}]{ticker} "
            f"({tier} {score:+d})[/]"
        )
        row1 = [
            row1_col1,
            f"{time_spotted_str}",
            f"{float_str}",
            f"[{style_for_rvol(rv_spot)}]{rv_spot:.1f}x[/]"
            if rv_spot is not None else "n/a",
            f"[{style_for_rsi(rsi_spot)}]{rsi_spot:.0f}[/]"
            if rsi_spot is not None else "n/a",
            f"${price_spot_str}",
            f"[{style_for_vol(vol_spot_str, float_shares)}]{vol_spot_str}[/]",
        ]

        # Row 2: G%spot->pk/G%alert->pk | TimeAlert | FR | RV | RSI | Price | Vol
        row2 = [
            f"{g_spot_to_peak:.1f}% / {g_alert_to_peak:.1f}%",
            f"{time_alerted_str}",
            f"[{style_for_fr(fr_alert)}]{fr_alert_str}[/]",
            f"[{style_for_rvol(rv_alert)}]{rv_alert:.1f}x[/]"
            if rv_alert is not None else "n/a",
            f"[{style_for_rsi(rsi_alert)}]{rsi_alert:.0f}[/]"
            if rsi_alert is not None else "n/a",
            f"${price_alert_str}",
            f"[{style_for_vol(vol_alert_str, float_shares)}]{vol_alert_str}[/]",
        ]

        # Row 3: AlertTime->PkTime | TimePk | FR | RV | RSI | Price | Vol
        row3 = [
            f"{time_to_peak_str}",
            f"{time_peaked_str}",
            f"[{style_for_fr(fr_peak)}]{fr_peak_str}[/]",
            f"[{style_for_rvol(rv_peak)}]"
            f"{rv_peak:.1f}x[/]" if rv_peak is not None else "n/a",
            f"[{style_for_rsi(rsi_peak)}]{rsi_peak:.0f}[/]"
            if rsi_peak is not None else "n/a",
            f"${price_peak_str}",
            f"{vol_peak_str}",
        ]

        table.add_row(*row1)
        table.add_row(*row2)
        table.add_row(*row3)

        # optional separator row between tickers
        table.add_row("", "", "", "", "", "", "")

    console.print(table)


def export_eod_stats_to_excel(
        winners: list[Stock], pct_chg_des: float, export_dir: Path
):
    """
    Create/write to an Excel file with the summary of stocks that met criteria.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None, "Logic error: openpyxl's wb.active is unexpectedly None!"
    ws.title = f"EOD Crit Stock Growth Summary"

    headers = [
        "CritTime",
        "CritPrice",
        "Ticker",
        "PeakTime",
        "MaxPrice",
        "Volume",
        "Peak%",
        "Score",
        "Tier",
        "Peak%FromSpot",
        "Time2Peak(min)",
        "FloatShares",
        "CritRVol",
        "CritRSI",
        "CritFR",
    ]
    ws.append(headers)

    # Format for date/time we want
    date_format = "HH:MM:SS"
    pct_format = "0.00%"
    currency_format = '"$"#,##0.00'

    # Store today's date/time for naming
    date_str = dt.now().strftime("%Y_%m_%d")

    # Append a row for each winner
    for s in winners:
        crit_time = s.get_crit_time()
        peak_time = s.get_time_max_price()
        # e.g. 20.0 => 0.2 in excel
        peak_change = s.get_peak_change() / 100.0
        peak_from_spot = s.get_peak_change_spot() / 100.0
        time_to_peak = s.get_time_peak_alert()

        row = [
            crit_time,  # datetime
            s.get_crit_price(),  # float
            s.get_ticker(),  # string
            peak_time,  # datetime
            s.get_max_price(),  # float
            s.get_vol_at_max_price(),  # string
            peak_change,  # float for actual %  e.g. 0.2 => 20%
            s.get_crit_score(),
            s.get_crit_tier(),
            peak_from_spot,
            time_to_peak,
            s.get_float_shares(),
            s.get_crit_rvol(),
            s.get_crit_rsi(),
            s.get_crit_vol_float_ratio(),
        ]
        ws.append(row)

    # Data starts at row 2
    for i in range(2, len(winners) + 2):
        # CritTime
        ws.cell(row=i, column=1).number_format = date_format
        # PeakTime
        ws.cell(row=i, column=4).number_format = date_format
        # CritPrice
        ws.cell(row=i, column=2).number_format = currency_format
        # MaxPrice
        ws.cell(row=i, column=5).number_format = currency_format
        # Peak%
        ws.cell(row=i, column=7).number_format = pct_format
        # Peak%FromSpot
        ws.cell(row=i, column=10).number_format = pct_format
        # 

    fname = f"eod_summary_{pct_chg_des}%_{date_str}.xlsx"
    export_dir.mkdir(parents=True, exist_ok=True)
    fpath = export_dir / fname

    wb.save(str(fpath))
    print(f"End-of-day summary exported to: {fpath}")


def export_all_seen_to_excel(stocks: list[Stock], pct_chg_des: float, export_dir: Path) -> None:
    """
    Export *every* ticker that appeared in Top Gainers at any point during the session.
    One row per ticker (summary-style, not tick-by-tick).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "All Top Gainers Seen"

    headers = [
        "TimeEntered",
        "Ticker",
        "OG_Price",
        "OG_Vol",
        "OG_RVol",
        "OG_RSI",
        "MetCriteria",
        "CritTime",
        "CritPrice",
        "CritScore",
        "Tier",
        "FloatShares",
        "CritRVol",
        "CritRSI",
        "CritFR",
        "PeakTime",
        "MaxPrice",
        "VolAtPeak",
        "Peak%FromSpot",
        "Peak%FromAlert",
        "TimeAlertToPeak(min)",
    ]
    ws.append(headers)

    date_format = "HH:MM:SS"
    pct_format = "0.00%"
    currency_format = '"$"#,##0.00'

    date_str = dt.now().strftime("%Y_%m_%d")

    for s in stocks:
        time_entered = s.get_time_entered()
        ticker = s.get_ticker()

        og_price = s.get_og_price()
        og_vol = s.get_og_vol()
        og_rvol = s.get_og_rvol()
        og_rsi = s.get_og_rsi()

        met = s.has_met_crit()
        crit_time = s.get_crit_time()
        crit_price = s.get_crit_price()
        crit_score = s.get_crit_score()
        crit_tier = s.get_crit_tier()

        float_shares = s.get_float_shares()
        crit_rvol = s.get_crit_rvol()
        crit_rsi = s.get_crit_rsi()
        crit_fr = s.get_crit_vol_float_ratio()

        peak_time = s.get_time_max_price()
        max_price = s.get_max_price()
        vol_at_peak = s.get_vol_at_max_price()

        # Excel wants percents as fractions (0.25 = 25%)
        peak_from_spot = s.get_peak_change_spot() / 100.0
        peak_from_alert = s.get_peak_change() / 100.0
        time_alert_to_peak = s.get_time_peak_alert()

        ws.append([
            time_entered,
            ticker,
            og_price,
            og_vol,
            og_rvol,
            og_rsi,
            met,
            crit_time,
            crit_price,
            crit_score,
            crit_tier,
            float_shares,
            crit_rvol,
            crit_rsi,
            crit_fr,
            peak_time,
            max_price,
            vol_at_peak,
            peak_from_spot,
            peak_from_alert,
            time_alert_to_peak,
        ])

    # Format time / currency / percent columns
    for i in range(2, len(stocks) + 2):
        ws.cell(row=i, column=1).number_format = date_format   # TimeEntered
        ws.cell(row=i, column=4).number_format = "@"           # OG_Vol (string)
        ws.cell(row=i, column=8).number_format = date_format   # CritTime
        ws.cell(row=i, column=16).number_format = date_format  # PeakTime

        ws.cell(row=i, column=3).number_format = currency_format   # OG_Price
        ws.cell(row=i, column=9).number_format = currency_format   # CritPrice
        ws.cell(row=i, column=17).number_format = currency_format  # MaxPrice

        ws.cell(row=i, column=19).number_format = pct_format   # Peak%FromSpot
        ws.cell(row=i, column=20).number_format = pct_format   # Peak%FromAlert

    fname = f"all_gainers_{pct_chg_des}%_{date_str}.xlsx"
    export_dir.mkdir(parents=True, exist_ok=True)
    fpath = export_dir / fname
    wb.save(str(fpath))
    print(f"All-seen gainers exported to: {fpath}")


def show_eod_stats(gainers: list[Stock], pct_chg_des: float):
    """
    End-of-day summary for all stocks that have met or surpassed pct_chg_des, in a Rich table.
    Saves to excel file if possible for record keeping.
    """
    winners = [s for s in gainers if s.has_met_crit()]
    if not winners:
        print("\nNo stocks met or surpassed your desired growth today.")
        return

    console = Console()
    table = Table(
        title=f"SUMMARY OF STOCKS THAT SURPASSED {pct_chg_des}% GROWTH TODAY",
        show_header=True,
        header_style="bold cyan",
    )

    table.add_column("CritTime")
    table.add_column("CritPrice", justify="right")
    table.add_column("Ticker", justify="left")
    table.add_column("PeakTime", justify="right")
    table.add_column("MaxPrice", justify="right")
    table.add_column("Volume", justify="right")
    table.add_column("Peak%", justify="right")
    table.add_column("Score", justify="right")
    table.add_column("Tier", justify="right")
    table.add_column("Peak%FromSpot", justify="right")
    table.add_column("Time2Peak", justify="right")
    table.add_column("FloatShares", justify="right")
    table.add_column("RVolume", justify="right")
    table.add_column("RSI", justify="right")
    table.add_column("FloatRatio", justify="right")

    for s in winners:
        crit_time = s.get_crit_time()
        time_max = s.get_time_max_price()
        assert crit_time is not None, "Logic error: we must have crit time here!"
        assert time_max is not None, "Logic error: we must have time max here!"
        crit_time_str = crit_time.strftime("%H:%M:%S")
        time_max_str = time_max.strftime("%H:%M:%S")
        peak_change = s.get_peak_change()
        peak_from_spot = s.get_peak_change_spot()
        time_to_peak = s.get_time_peak_alert()
        crit_score = s.get_crit_score() if s.get_crit_score() is not None else 0
        float_shares = s.get_float_shares()
        crit_rvol = s.get_crit_rvol()
        crit_rsi = s.get_crit_rsi()
        crit_vol_float_ratio = s.get_crit_vol_float_ratio()
        crit_vol_float_ratio = round(crit_vol_float_ratio, 2) if crit_vol_float_ratio is not None else None


        volume_str = s.get_vol_at_max_price()
        peak_str = colorize_pct(peak_change)
        crit_price_str = f"${s.get_crit_price():.2f}"
        max_price_str = f"${s.get_max_price():.2f}"
        peak_from_spot_str = colorize_pct(peak_from_spot)
        time_to_peak_str = f"{time_to_peak} mins" if time_to_peak is not None else "--"
        crit_score_str = f"{crit_score:+d}"
        crit_tier_str = s.get_crit_tier() if s.get_crit_tier() is not None else "--"
        float_shares_str = millify(float_shares) if float_shares else "--"
        crit_rvol_str = str(crit_rvol)
        crit_rsi_str = str(crit_rsi)
        crit_vol_float_ratio_str = str(crit_vol_float_ratio)

        table.add_row(
            crit_time_str,
            crit_price_str,
            s.get_ticker(),
            time_max_str,
            max_price_str,
            volume_str,
            peak_str,
            crit_score_str,
            crit_tier_str,
            peak_from_spot_str,
            time_to_peak_str,
            float_shares_str,
            crit_rvol_str,
            crit_rsi_str,
            crit_vol_float_ratio_str
        )

    console.print(table)
    if EXPORT_PATH is not None:
        export_eod_stats_to_excel(winners, pct_chg_des, EXPORT_PATH)
    else:
        print("No EOD_EXPORT_PATH environment variable found; skipping Excel export.")


def run_main_loop(
    MARKET_CLOSE: dt,
    driver: webdriver.Chrome,
    float_prov: FloatProvider,
    gainers: list[Stock],
    ref_rate_des: float,
    pct_chg_des: float,
):
    """
    The main loop that repeatedly scrapes the gainers table, updates stocks,
    checks user criteria, displays top 5 if changed, etc.
    """
    alerted: list[Stock] = []

    while dt.now() < MARKET_CLOSE:
        # 1) Refresh the browser to get latest data
        try:
            driver.refresh()
        except TimeoutException:
            print("\n\033[1;33m[WARNING]\033[0m Page refresh timed out, continuing anyway...")
        try:
            wait_for_table(driver, 12)
        except TimeoutException:
            print("\n\033[1;33m[WARNING]\033[0m Table did not fully load after refresh...")
  
        # 2) Scrape
        overview_rows = scrape_overview(driver)
        technicals_map = scrape_technicals(driver)

        new_data = []
        for row in overview_rows:
            rsi = technicals_map.get(row["ticker"])
            new_data.append(
                (row["ticker"], row["price"], row["vol"], row["rvol"], rsi)
            )
        # 3) Process
        changed = process_stocks(
            gainers, float_prov, alerted, new_data, pct_chg_des
        )
        # 4) If changed, sort and show top 5
        if changed:
            gainers.sort(key=lambda s: s.get_abs(), reverse=True)
        # 5) Sleep
        sleep(ref_rate_des if changed else 10)


def main():
    """
    Main entry point for the StockBot application.
    Checks market day, waits if before open, obtains user params, logs in, and runs main loop.
    """
    float_prov = FloatProvider()
    ref_rate_des, pct_chg_des = get_user_params()
    # Setup driver & login
    driver = setup_webdriver()
    try:
        driver.get(URL)
        logged_in = is_logged_in(driver)
        print(
                "You",
                "\033[1;32mare\033[0m logged in." if logged_in
                        else "\033[1;31mare not\033[0m logged in."
        )
        if not logged_in:
            if HEADLESS:
                raise RuntimeError(
                        "Not logged in. Run once non-headless to authenticate.")
            driver.get(URL)
            login_to_tradingview(driver, EMAIL, PASS)
            # Check recaptcha
            check_recaptcha(driver)
            # Make sure to wait for login (reCAPTCHA can be tricky with this)
            wait_for_login(driver)

        # Go to Gainers page
        driver.get(GAINS_URL)

        while True:
            # Get next market day
            now = dt.now()
            next_open, next_close = get_next_day(now)
            sec_until_open = (next_open - now).total_seconds() 
            # Create main list
            gainers = []
            # Wait until market open of next available day
            if sec_until_open > 0:
                print(f"\nWaiting until market open on {next_open.strftime('%a @ %H:%M:%S')}...")
                sleep(sec_until_open)
            # Run main loop
            run_main_loop(
                next_close, driver, float_prov, gainers, ref_rate_des, pct_chg_des
            )

            # End-of-day summary
            show_eod_stats(gainers, pct_chg_des)
            # Export everything we saw even if no stocks hit criteria
            if EXPORT_PATH is not None:
                export_all_seen_to_excel(gainers, pct_chg_des, EXPORT_PATH)

            print("\nMarket CLOSED now!\n")
    except KeyboardInterrupt:
        print("\nEnding program...")
    finally:
        try:
            driver.quit()
        except Exception:
            pass


if __name__ == "__main__":
    main()
